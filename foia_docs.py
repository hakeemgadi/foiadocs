import os
from selenium import webdriver
import datetime
import time
import urllib.request
import openpyxl
import pickle

date_format = '%Y-%m-%d'

#Function to get date from user
def get_date():
        while True:
                date=input("Please enter the date in the YYYY-MM-DD format\n")

                try:
                        date_obj = datetime.datetime.strptime(date, date_format)
                        break
                except ValueError:
                        print("Incorrect data format, should be YYYY-MM-DD. Please try again\n")
        return date_obj.strftime('%Y-%m-%d')

#A class for fields in a row
class Field(object):
        def __init__(self,text=None,link=None):
                self.text=text
                self.link=link
                
        def __str__(self):
                return ' '.join([str(self.text), str(self.link)])

#A class for a single data point (To facilitate handling data points)
class DataPoint(object):
        def __init__(self):
                self.entry={'subject':Field(),
                            'local_link':Field('Local link'),
                            'foia_link':Field('FOIA link'),
                            'file_name':Field(),
                            'document_date':Field(),
                            'from':Field(),
                            'to':Field(),
                            'posted_date':Field(),
                            'case_no':Field()}
        def __str__(self):
                accumulator=[]
                for key,val in self.entry.items():
                        accumulator.append(' '.join([key+':',str(val.text),str(val.link)]))
                return '\n'.join(accumulator)
        
        def __getitem__(self,index):
                if index >= len(self.entry):
                        raise StopIteration
                return list(self.entry.values())[index]

#A subclass of openpyxl.Workbook (To populate the active worksheet from a data point)
class FOIA_Workbook(openpyxl.Workbook):
        def update_row_from_dpoint(self,dpoint,row):
                for field in dpoint:
                        self.active.cell(row,i+1).value= field.text
                        if field.link is not None:
                                self.active.cell(row,i+1).hyperlink= field.link
                

#Initialise Chrome webdrive        
option = webdriver.ChromeOptions()
chrome_prefs = {}
option.experimental_options["prefs"] = chrome_prefs
#Disable image download
chrome_prefs["profile.default_content_settings"] = {"images": 2}
chrome_prefs["profile.managed_default_content_settings"] = {"images": 2}
driver = webdriver.Chrome(options=option,executable_path=os.path.join("..","..","chromedriver.exe"))

#Get input from user. Search string and date range
search_string = input('Please enter the search term:\n')
print('Please enter the search begin date:\n')
beginDate = get_date() #yyyymmdd
print('Please enter the search end date: \n')
endDate= get_date()

print(f"Downloading files for search term [{search_string}] beginning in {beginDate} and ending in {endDate}")

#Navigate to the application home page
driver.get(f"https://foia.state.gov/Search/Results.aspx?searchText={search_string}&beginDate={beginDate}&endDate={endDate}&publishedBeginDate=&publishedEndDate=&caseNumber=")

#Create workbook to store tabulated data
wb = FOIA_Workbook()
#ws = wb.active



time.sleep(2)
page_sz=20

startpg=1 #Normally 1. Unless not starting in the middle


#Get end page number by reading the number of pages
end_pg=int(driver.find_element_by_id("lblPagesTop").text)

#Loop pages
for i in range(startpg-1,end_pg):
        
        #Get result table in page
        result_table=driver.find_element_by_id("tblResults")

        #Get all rows in table
        rows=result_table.find_elements_by_tag_name("tr")

        #Loop result table rows
        for j,row in enumerate(rows):

                #dpoint will store the fields in the current row. It is serialized later to serve as a backup for Excel rows.
                dpoint=DataPoint()

                #Get all fields in row
                data=row.find_elements_by_tag_name("td")

                #Populate Excel workbook 
                if len(data)!=0:
                        #Write text into row i*page_sz+j
                        cell_row = i*page_sz+j
                        
                        print(f"Processing row no.: {cell_row}")

                        #Subject column
                        dpoint.entry['subject'].text = data[1].text
                        

                        #Get pdf link from subject column
                        pdflink = data[1].find_element_by_tag_name("span").get_attribute("title")

                        #Get file name only
                        file = pdflink.split("/")[-1]

                        #Construct local path of pdf
                        dpoint.entry['local_link'].link = os.path.join("..","DocStore","PDFS",file)

                        
                        #Construct a global weblink from site-local link
                        dpoint.entry['foia_link'].link = "https://foia.state.gov/"+pdflink


                        dpoint.entry['file_name'].text = file

                        #Process the remaining of fields: Document Date, From, To, Posted Date, and Case Number
                        for rest in range(4,9):
                                #Offset by 3 columns in the Excel workbook
                                dpoint[rest].text = data[rest-2].text
                        print(dpoint)
                        wb.update_row_from_dpoint(dpoint,cell_row)
                        
                        #Download the PDF
                        urllib.request.urlretrieve("https://foia.state.gov/"+pdflink, os.path.join("..","DocStore","PDFS",file))

                        dfile=open(os.path.join("..","DocStore","PKL",str(cell_row)+".pkl"),"wb")
                        pickle.dump(dpoint,dfile)
                        dfile.close()
                        print("=========")

                        
                        time.sleep(2)
        if i < end_pg-1:
                #Go to next page if not last page

                inputElement = driver.find_element_by_id("txtPageTop")
                inputElement.clear()
                inputElement.send_keys(i+2)

                btn=driver.find_element_by_id("btnJumpTop")

                btn.click()
                print("Going to next page ...")

        time.sleep(3)

#Save workbook
wb.save(os.path.join("..","DocStore",'FilesLog.xlsx')) 

